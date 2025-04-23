import os 
from pathlib import Path
import streamlit as st
import cv2
import numpy as np
from deepface import DeepFace
import mediapipe as mp
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import pandas as pd
import uuid
from datetime import datetime
import base64
from PIL import Image as PILImage
import io
import time

# Initialize directories
BASE_DIR = Path(__file__).parent
CAPTURED_FACES_DIR = BASE_DIR / "captured_faces"
REPORTS_DIR = BASE_DIR / "reports"

CAPTURED_FACES_DIR.mkdir(exist_ok=True)
REPORTS_DIR.mkdir(exist_ok=True)

# Initialize MediaPipe
mp_face_detection = mp.solutions.face_detection
face_detection = mp_face_detection.FaceDetection(min_detection_confidence=0.5)

# Session storage
if 'sessions' not in st.session_state:
    st.session_state.sessions = {}
if 'current_session' not in st.session_state:
    st.session_state.current_session = None
if 'emotion_logs' not in st.session_state:
    st.session_state.emotion_logs = {}
if 'participants' not in st.session_state:
    st.session_state.participants = {}

def generate_meeting_code():
    return str(uuid.uuid4())[:8]

def capture_frame(video_capture):
    ret, frame = video_capture.read()
    if not ret:
        return None
    return cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

def detect_emotion(frame):
    try:
        result = DeepFace.analyze(frame, actions=['emotion'], enforce_detection=False)
        return result[0]['dominant_emotion']
    except:
        return "neutral"

def detect_face(frame):
    results = face_detection.process(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
    if results.detections:
        return True
    return False

def generate_excel_report(session_id, participants, emotion_logs):
    wb = Workbook()
    ws = wb.active
    ws.title = "Session Report"

    headers = ["Student ID", "Name", "Dominant Emotion", "Photo"]
    for col, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col)}1"] = header

    for row, (student_id, data) in enumerate(participants.items(), 2):
        ws[f"A{row}"] = student_id
        ws[f"B{row}"] = data.get('name', 'Unknown')
        
        # Calculate dominant emotion
        emotions = [log['emotion'] for log in emotion_logs.get(student_id, [])]
        dominant_emotion = max(set(emotions), key=emotions.count) if emotions else "N/A"
        ws[f"C{row}"] = dominant_emotion

        # Add photo if exists
        photo_path = data.get('photo_path')
        if photo_path and os.path.exists(photo_path):
            img = PILImage.open(photo_path)
            img = img.resize((100, 100))
            img_buffer = io.BytesIO()
            img.save(img_buffer, format="PNG")
            
            img_obj = Image(img_buffer)
            ws.add_image(img_obj, f"D{row}")
            ws.row_dimensions[row].height = 100
            ws.column_dimensions['D'].width = 20

    report_path = REPORTS_DIR / f"session_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(report_path)
    return report_path

def get_binary_file_downloader_html(file_path, file_name):
    with open(file_path, "rb") as f:
        data = f.read()
    b64 = base64.b64encode(data).decode()
    href = f'<a href="data:application/octet-stream;base64,{b64}" download="{file_name}">Download {file_name}</a>'
    return href

def teacher_dashboard():
    st.header("Teacher Dashboard")
    
    if st.button("Create New Session"):
        session_id = generate_meeting_code()
        st.session_state.sessions[session_id] = {
            'active': True,
            'recording': False,
            'start_time': datetime.now()
        }
        st.session_state.current_session = session_id
        st.success(f"New session created with code: {session_id}")

    if st.session_state.current_session:
        session_id = st.session_state.current_session
        st.write(f"Current Session Code: {session_id}")
        
        # Screen sharing placeholder (using iframe for demonstration)
        st.subheader("Screen Sharing")
        st.write("Screen sharing functionality would be implemented using WebRTC")
        
        # Recording controls
        if st.button("Start/Stop Recording"):
            st.session_state.sessions[session_id]['recording'] = not st.session_state.sessions[session_id].get('recording', False)
        
        if st.session_state.sessions[session_id].get('recording'):
            st.write("Recording in progress...")
        
        # Display participants
        st.subheader("Participants")
        for student_id, data in st.session_state.participants.items():
            col1, col2, col3 = st.columns(3)
            with col1:
                st.write(f"ID: {student_id}")
            with col2:
                st.write(f"Name: {data.get('name', 'Unknown')}")
            with col3:
                emotions = [log['emotion'] for log in st.session_state.emotion_logs.get(student_id, [])]
                dominant_emotion = max(set(emotions), key=emotions.count) if emotions else "N/A"
                st.write(f"Emotion: {dominant_emotion}")

        # Generate and download report
        if st.button("Generate Report"):
            report_path = generate_excel_report(session_id, st.session_state.participants, st.session_state.emotion_logs)
            st.markdown(get_binary_file_downloader_html(report_path, os.path.basename(report_path)), unsafe_allow_html=True)

def student_dashboard():
    st.header("Student Dashboard")
    
    session_code = st.text_input("Enter Session Code")
    student_name = st.text_input("Enter Your Name")
    
    if st.button("Join Session"):
        if session_code in st.session_state.sessions:
            student_id = str(uuid.uuid4())[:8]
            st.session_state.participants[student_id] = {'name': student_name}
            st.session_state.current_session = session_code
            st.success("Joined session successfully!")
        else:
            st.error("Invalid session code")

    if st.session_state.current_session:
        # Webcam feed
        video_capture = cv2.VideoCapture(0)
        frame_placeholder = st.empty()
        
        while True:
            frame = capture_frame(video_capture)
            if frame is None:
                break
                
            # Face detection and emotion analysis
            if detect_face(frame):
                emotion = detect_emotion(frame)
                
                # Save face capture
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                face_path = CAPTURED_FACES_DIR / f"student_{student_id}_{timestamp}.png"
                cv2.imwrite(str(face_path), cv2.cvtColor(frame, cv2.COLOR_RGB2BGR))
                
                # Log emotion
                if student_id not in st.session_state.emotion_logs:
                    st.session_state.emotion_logs[student_id] = []
                st.session_state.emotion_logs[student_id].append({
                    'timestamp': datetime.now(),
                    'emotion': emotion
                })
                
                st.session_state.participants[student_id]['photo_path'] = str(face_path)
                
                # Display frame with emotion overlay
                cv2.putText(frame, f"Emotion: {emotion}", (10, 30), 
                           cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
            
            frame_placeholder.image(frame, channels="RGB")
            time.sleep(0.1)
        
        video_capture.release()

def main():
    st.title("Virtual Classroom Platform")
    
    role = st.selectbox("Select Role", ["Teacher", "Student"])
    
    if role == "Teacher":
        teacher_dashboard()
    else:
        student_dashboard()

if __name__ == "__main__":
    main()